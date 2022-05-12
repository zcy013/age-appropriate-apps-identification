#!/usr/bin/python3
# -*- coding: UTF-8 -*-
# author: zcy

import os
import time
import operator
import pandas as pd
import numpy as np
import openpyxl
import itertools

from sklearn.model_selection import cross_val_score  # 交叉验证
from sklearn.linear_model import LogisticRegression
from sklearn.svm import SVC
from sklearn.tree import DecisionTreeClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.neural_network import MLPClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.feature_selection import RFE, SelectKBest, SelectPercentile, chi2
from sklearn.model_selection import GridSearchCV
from sklearn.metrics import accuracy_score
from sklearn.pipeline import make_pipeline
from lime.lime_text import LimeTextExplainer
from lime.lime_tabular import LimeTabularExplainer


from configs import random_state, MODEL_IMGS, MODEL_OCRS, feature_size_assigned
from utils import read_apk_features, get_all_feature_files_and_labels, get_part_feature_files_and_labels, train_models_grid, train_default_models
import learner_text


# 统计data中每个子数据的出现次数，累计到counter
def count_sub_data(counter, data):
    for k, v in data.items():
        if k in counter:
            counter[str(k)] += v
        else:
            counter[str(k)] = v


# 根据总数量，取前N%
def top_data_filter(dicts, total_count, filter_ratio):
    sort_dict = dict(sorted(dicts.items(), key=operator.itemgetter(1), reverse=True))
    sort_dict = dict((key, value) for key, value in sort_dict.items() if value > total_count * filter_ratio)
    return list(sort_dict.keys())


# 统计apk的指定feature的指定子feature
def sub_features_encoder(feature_map, feature_name, img):
    sub = []
    for f in feature_map[feature_name]:
        status = img[feature_name][f] if f in img[feature_name] else 0
        sub.append(status)
    return sub


# 为指定子feature赋予定好的权重
def sub_features_encoder_weighted(feature_map, feature_name, img):
    sub = []
    for weight, labels in feature_map.items():
        for f in labels:
            status = img[feature_name][f] * weight if f in img[feature_name] else 0
            sub.append(status)
    return sub


class MyGallery:
    def __init__(self, file_list=None, labels=None, sub_features=None):
        # 指定apk文件路径
        if file_list:
            self.file_list = file_list
            self.labels = labels
        else:
            # self.file_list, self.labels = get_part_feature_files_and_labels(1000, 1000)
            self.file_list, self.labels = get_all_feature_files_and_labels()
        self.file_list = [f.replace('features_all', 'features_img') for f in self.file_list]
        if self.labels:
            self.pos_count = self.labels.count(1)
            self.neg_count = len(self.labels) - self.pos_count
        if sub_features:
            self.sub_features = sub_features
        else:
            self.sub_features = {'avg_hsv': ['avg_h', 'avg_s', 'avg_v']}
        self.weights = {}
        self.dataframe_keys = []
        self.library = []

    # 1. 对规模大的feature，建立一个子feature映射，只考虑在映射中的子feature
    def build_sub_feature_map(self, ratio, labels_adt=None, labels_kid=None):
        self.sub_features['safe'] = ['adult', 'medical', 'spoofed', 'violence', 'racy']
        # self.sub_features['objects'] = self.get_popular_features_by_files('objects', ratio)

        # 读取label的人工标注
        if labels_adt and labels_kid:
            # labels_adt, labels_kid = load_image_labels()
            self.sub_features['labels'] = labels_adt + labels_kid
            self.weights = {-1: labels_adt, 1: labels_kid}

        # 统计某个feature中每种子feature的出现频率，之后正式统计时只考虑这些出现频率多的
        else:
            self.sub_features['labels'] = self.get_popular_features_by_files('labels', ratio)  # comment
        # self.sub_features['avg_hsv'] = ['avg_h', 'avg_s', 'avg_v']
        for k, v in self.sub_features.items():
            self.dataframe_keys += v

    def get_popular_features_by_files(self, feature_name, ratio):
        # 正例负例分别统计子特征，取出现频率前N%的子特征，然后取交集
        f_kid, f_adt = self.count_sub_features_all(feature_name)
        # tmp 统计在正负例中的得分
        # with open('image_labels.xlsx', 'w') as file:
        #     file.write('image label\tadt value\tkid value\n')
        #     for feature, adt_value in f_adt.items():
        #         kid_value = f_kid.pop(feature, 0)
        #         file.write('%s\t%.2f\t%.2f\n' % (feature, adt_value, kid_value))
        #     for feature, kid_value in f_kid.items():
        #         file.write('%s\t%.2f\t%.2f\n' % (feature, 0, kid_value))
        f_kid = top_data_filter(f_kid, self.pos_count, ratio)
        f_adt = top_data_filter(f_adt, self.neg_count, ratio)
        final_features = list(set(f_kid).union(set(f_adt)))
        # final_features = [x for x in f_kid if x in f_adt]
        # print('[%s] counted feature: %s\n%s' % (time.strftime('%Y-%m-%d %H:%M:%S'), feature_name, final_features))
        return final_features

    def count_sub_features_all(self, feature_name):
        feature_count_kid = {}
        feature_count_adt = {}
        for i in range(len(self.file_list)):
            path = self.file_list[i]
            imgs = read_apk_features(path)
            if self.labels[i]:
                for img_title, img_features in imgs.items():
                    count_sub_data(feature_count_kid, img_features[feature_name])
            else:
                for img_title, img_features in imgs.items():
                    count_sub_data(feature_count_adt, img_features[feature_name])
        return feature_count_kid, feature_count_adt

    # ocr的处理，此feature分类尴尬故单独处理
    def get_all_strings(self):
        if not self.library:
            for path in self.file_list:
                imgs = read_apk_features(path)
                ocrs = []
                for img, features in imgs.items():
                    if 'string_en' in features:
                        ocrs.append(features['string_en'])
                    else:
                        ocrs.append(features['string'])
                if not ' '.join(ocrs).strip():
                    print(path, ocrs)
                    self.library.append('.')
                else:
                    self.library.append(' '.join(ocrs))
        return self.library

    # 2.
    # 图片feature转应用feature 方案1 所有feature的平均值
    def convert_files_to_dataframe_avg(self, feature_names=None, to_scale=False):
        if not feature_names:
            feature_names = ['labels', 'safe']
        dataframe_keys = []
        for feature_name in feature_names:
            if feature_name in self.sub_features.keys():
                dataframe_keys += self.sub_features[feature_name]
            else:
                dataframe_keys.append(feature_name)
        if not self.weights:
            ladt, lkid = load_best_image_labels()
            self.weights = {-1: ladt, 1: lkid}
        all_apks_features = []
        for f in self.file_list:
            apk_features = [0 for i in dataframe_keys]
            # print('[%s] reading file %s' % (time.strftime('%Y-%m-%d %H:%M:%S'), f))
            imgs = read_apk_features(f)
            for img_title, img_features in imgs.items():
                # 处理当前应用下每个图片
                img_features_tuned = []
                for feature_name, sub_features in img_features.items():
                    if feature_name in feature_names:
                        if feature_name in self.sub_features.keys():
                            # if feature_name == 'labels':
                            #     img_features_tuned += sub_features_encoder_weighted(self.weights, feature_name, img_features)
                            # else:
                            img_features_tuned += sub_features_encoder(self.sub_features, feature_name, img_features)
                        else:
                            img_features_tuned.append(img_features[feature_name])
                # 图片转应用
                # 先求和
                # print(len(apk_features), len(img_features_tuned), len(dataframe_keys))
                apk_features = [apk_features[i] + img_features_tuned[i] for i in range(len(dataframe_keys))]
            # 再取平均
            num_imgs = len(imgs.items())
            apk_features = [apk_features[i] / num_imgs for i in range(len(dataframe_keys))]

            all_apks_features.append(apk_features)
        return pd.DataFrame(all_apks_features, columns=dataframe_keys)

    # 图片feature转应用feature 方案2 每个feature取最大值
    def convert_files_to_dataframe_max(self, feature_names=None, to_scale=False):
        if not feature_names:
            feature_names = ['labels', 'safe']
        dataframe_keys = []
        for feature_name in feature_names:
            if feature_name in self.sub_features.keys():
                dataframe_keys += self.sub_features[feature_name]
            else:
                dataframe_keys.append(feature_name)
        all_apks_features = []

        for f in self.file_list:
            apk_features = [0 for i in dataframe_keys]
            # print('[%s] reading file %s' % (time.strftime('%Y-%m-%d %H:%M:%S'), f))
            imgs = read_apk_features(f)
            for img_title, img_features in imgs.items():
                # 处理当前应用下每个图片
                img_features_tuned = []
                for feature_name, sub_features in img_features.items():
                    if feature_name in feature_names:
                        if feature_name in self.sub_features.keys():
                            if feature_name == 'labels':
                                img_features_tuned += sub_features_encoder_weighted(self.weights, feature_name, img_features)
                            else:
                                img_features_tuned += sub_features_encoder(self.sub_features, feature_name, img_features)
                        else:
                            img_features_tuned.append(img_features[feature_name])
                # 图片转应用
                # apk_features = [max(apk_features[i], img_features_tuned[i]) for i in range(len(dataframe_keys))]
                for i in range(len(dataframe_keys)):
                    if abs(apk_features[i]) < abs(img_features_tuned[i]):
                        apk_features[i] = img_features_tuned[i]

            all_apks_features.append(apk_features)
        return pd.DataFrame(all_apks_features, columns=dataframe_keys)

    # 图片feature转应用feature 方案3 只记录是否有每个feature 效果差，不再用
    # def convert_files_to_dataframe_0_1(self, feature_names=None):
    #     if not feature_names:
    #         feature_names = ['labels', 'safe']
    #     dataframe_keys = []
    #     for feature_name in feature_names:
    #         dataframe_keys += self.sub_features[feature_name]
    #     all_apks_features = []
    #
    #     for f in self.file_list:
    #         apk_features = [0 for i in dataframe_keys]
    #         # print('[%s] reading file %s' % (time.strftime('%Y-%m-%d %H:%M:%S'), f))
    #         imgs = read_apk_features(f)
    #         counter = 0
    #         for img_title, img_features in imgs.items():
    #             counter += 1
    #             # 处理当前应用下每个图片
    #             img_features_tuned = []
    #             for feature_name, sub_features in img_features.items():
    #                 if feature_name in feature_names:
    #                     img_features_tuned += sub_features_encoder(self.sub_features, feature_name, img_features)
    #             # 图片转应用
    #             apk_features = [1 if img_features_tuned[i] > 0 else 0 for i in range(len(dataframe_keys))]
    #             if counter >= 10:
    #                 break
    #         all_apks_features.append(apk_features)
    #     return pd.DataFrame(all_apks_features, columns=dataframe_keys)


def load_all_image_labels():
    col_cy, col_ll, col_zx = 8, 10, 12
    wb = openpyxl.load_workbook('image_labels.xlsx')
    ws = wb.worksheets[0]

    labels_adt1, labels_kid1 = [], []
    labels_adt2, labels_kid2 = [], []
    labels_adt3, labels_kid3 = [], []
    for row in ws.iter_rows():
        for col, ladt, lkid in zip([col_cy, col_ll, col_zx],
                                   [labels_adt1, labels_adt2, labels_adt3],
                                   [labels_kid1, labels_kid2, labels_kid3]):
            if row[col].value == -1:
                ladt.append(row[0].value)
            elif row[col].value == 1:
                lkid.append(row[0].value)

    # 单独
    print('cy')
    yield labels_adt1, labels_kid1
    print('ll')
    yield labels_adt2, labels_kid2
    print('zx')
    yield labels_adt3, labels_kid3

    models_part = list(itertools.combinations([1, 2, 3], 2))
    for com in models_part:
        to_combine_adt, to_combine_kid = [], []
        if 1 in com:
            to_combine_adt.append(labels_adt1)
            to_combine_kid.append(labels_kid1)
        if 2 in com:
            to_combine_adt.append(labels_adt2)
            to_combine_kid.append(labels_kid2)
        if 3 in com:
            to_combine_adt.append(labels_adt3)
            to_combine_kid.append(labels_kid3)
        # 两两相交
        print('intersec of 2')
        yield list(set(to_combine_adt[0]).intersection(set(to_combine_adt[1]))), \
              list(set(to_combine_kid[0]).intersection(set(to_combine_kid[1]))),
        # 两两相并
        print('union of 2')
        yield list(set(to_combine_adt[0]).union(set(to_combine_adt[1]))), \
              list(set(to_combine_kid[0]).union(set(to_combine_kid[1]))),

    # 三者相交
    print('intersec of 3')
    yield list(set(labels_adt1).intersection(set(labels_adt2).intersection(set(labels_adt3)))), \
          list(set(labels_kid1).intersection(set(labels_kid2).intersection(set(labels_kid3)))),
    # 三者相并
    print('union of 3')
    yield list(set(labels_adt1).union(set(labels_adt2).union(set(labels_adt3)))), \
          list(set(labels_kid1).union(set(labels_kid2).union(set(labels_kid3)))),

    # 三者多数投票
    major_adt, major_kid = [], []
    for row in ws.iter_rows():
        counter_adt, counter_kid = 0, 0
        for col in [col_cy, col_ll, col_zx]:
            if row[col].value == -1:
                counter_adt += 1
            elif row[col].value == 1:
                counter_kid += 1
        if counter_adt >= 2:
            major_adt.append(row[0].value)
        elif counter_kid >= 2:
            major_kid.append(row[0].value)
    print('major')
    yield major_adt, major_kid
    return


def load_best_image_labels():
    col_cy, col_ll, col_zx = 8, 10, 12
    wb = openpyxl.load_workbook('image_labels.xlsx')
    ws = wb.worksheets[0]
    ladt, lkid = [], []
    for row in ws.iter_rows():
        if row[col_ll].value == -1 or row[col_zx].value == -1:
            ladt.append(row[0].value)
        elif row[col_ll].value == 1 or row[col_zx].value == 1:
            lkid.append(row[0].value)
    return ladt, lkid


def hard_classification(X, counter, bound):
    Y = []
    for idx, data in X.iterrows():
        sum = 0
        for v in data:
            if v <= bound:
                sum += 1
        # print(sum)
        if sum >= counter:
            Y.append(0)
        else:
            Y.append(1)
    return Y


def train_hard_grid(X, Y):
    best_acc = 0
    best_counter = 0
    best_bound = 0
    for counter in range(1, 26):
        for bound in range(-10, 0):
            acc = accuracy_score(hard_classification(X, counter, bound / 10), Y)
            if acc > best_acc:
                best_acc = acc
                best_counter = counter
                best_bound = bound / 10
    print(best_acc, best_counter, best_bound)
    return best_acc, best_counter, best_bound


def select_best_chi2(X, Y, ratio, size=0):
    if size:
        model = SelectKBest(chi2, k=size)
    else:
        model = SelectPercentile(chi2, percentile=ratio)
    X_new = model.fit_transform(X, Y)
    # print(model.scores_)  # 得分越高，特征越重要
    # print(model.pvalues_)  # p-values 越小，置信度越高，特征越重要
    # mask = model.get_support()  # 得到每个特征是否被选择的T-F数组
    # indices = model.get_support(indices=True)  # 得到所有选择出的特征的索引数组（按索引顺序）
    indices = np.argsort(model.scores_)[::-1]  # 得到所有选择出的特征的索引数组（按score倒序）
    # best_features = list(X.columns.values[indices[0:feature_size]])
    for i in range(X_new.shape[1]):
        index = indices[i]
        if model.scores_[index] == 0:
            break
        print('%d\t%s\t%f\t%f' % (i+1, X.columns.values[index], model.scores_[index], model.pvalues_[index]))
    X_new = pd.DataFrame(X_new, columns=X.columns.values[model.get_support(indices=True)])
    return X_new


# def explain_model(name, classifier, file, label):
#     imgs = read_apk_features(file.replace('features_all', 'features_img'))
#     if name == MODEL_OCRS:
#         # 文本类，合在一起
#         ocrs = []
#         for img, features in imgs.items():
#             ocrs.append(features['string'])
#         if not ' '.join(ocrs).strip():
#             print(file, ocrs)
#             data = '。'
#         else:
#             data = '。'.join(ocrs)
#
#         # vector = bert_client.encode(data)
#         pipeline = make_pipeline(learner_text.BertVectorizer(learner_text.bert_client), classifier)
#
#         explainer = LimeTextExplainer(class_names=['adt', 'kid'])
#         exp = explainer.explain_instance(data, pipeline.predict_proba, num_features=6)
#         print('true label =', ['adt', 'kid'][label])
#         print('prob(kid) =', pipeline.predict_proba([data])[0, 1])
#         print(exp.as_list())
#         # exp.save_to_file('exp.html', text=True)
#     elif name == MODEL_IMGS:
#         pass


def train_models(X, Y):
    print('[%s] Decision Tree:' % time.strftime('%Y-%m-%d %H:%M:%S'))
    model = DecisionTreeClassifier()
    scores = cross_val_score(model, X, Y, cv=10)
    print(scores)
    print('avg = %f' % np.mean(scores))

    print('[%s] Logistic Regression:' % time.strftime('%Y-%m-%d %H:%M:%S'))
    model = LogisticRegression()
    scores = cross_val_score(model, X, Y, cv=10)
    print(scores)
    print('avg = %f' % np.mean(scores))

    print('[%s] Linear SVM:' % time.strftime('%Y-%m-%d %H:%M:%S'))
    model = SVC()
    scores = cross_val_score(model, X, Y, cv=10)
    print(scores)
    print('avg = %f' % np.mean(scores))

    print('[%s] K-Nearest Neighbors:' % time.strftime('%Y-%m-%d %H:%M:%S'))
    model = KNeighborsClassifier()
    scores = cross_val_score(model, X, Y, cv=10)
    print(scores)
    print('avg = %f' % np.mean(scores))

    print('[%s] Random Forest:' % time.strftime('%Y-%m-%d %H:%M:%S'))
    model = RandomForestClassifier()
    scores = cross_val_score(model, X, Y, cv=10)
    print(scores)
    print('avg = %f' % np.mean(scores))

    print('[%s] Multi-Layer Perceptron:' % time.strftime('%Y-%m-%d %H:%M:%S'))
    model = MLPClassifier()
    scores = cross_val_score(model, X, Y, cv=10)
    print(scores)
    print('avg = %f' % np.mean(scores))


def get_trained_model(model_name, file_list, labels_true):
    gallery = MyGallery(file_list, labels_true)
    print('kid:', gallery.pos_count, 'adt', gallery.neg_count)
    if model_name == 'image':
        print('\n[%s] counting features ...' % time.strftime('%Y-%m-%d %H:%M:%S'))
        # labels_adt, labels_kid = load_best_image_labels()
        gallery.build_sub_feature_map(0)  # 0.3指取前70%
        X = gallery.convert_files_to_dataframe_avg(['labels'])
        X_filtered = select_best_chi2(X, labels_true, 0, feature_size_assigned[model_name])
        model = SVC(probability=True)
        model.fit(X_filtered, labels_true)
        return model, {'labels': list(X_filtered.columns.values)}
    elif model_name == 'ocrs':
        X = learner_text.bert_client.encode(gallery.get_all_strings())
        model = RandomForestClassifier()
        model.fit(X, labels_true)
        return model
    else:  # hsv
        gallery.build_sub_feature_map(0)
        X = gallery.convert_files_to_dataframe_avg(['avg_hsv'])
        model = SVC(probability=True)
        model.fit(X, labels_true)
        return model


def predict_proba(model_name, file_list, model, args=None):
    if args:
        gallery = MyGallery(file_list, sub_features=args[0])
    else:
        gallery = MyGallery(file_list)
    if model_name == 'image':
        X = gallery.convert_files_to_dataframe_avg(['labels'])
    elif model_name == 'ocrs':
        X = learner_text.bert_client.encode(gallery.get_all_strings())
    else:
        X = gallery.convert_files_to_dataframe_avg(['avg_hsv'])
    return model.predict_proba(X)


def train_callee(name, file_list, labels, grid):
    gallery = MyGallery(file_list, labels)
    print('kid:', gallery.pos_count, 'adt', gallery.neg_count)
    Y = gallery.labels
    if name == MODEL_IMGS:
        # for img features
        print('\n[%s] counting features ...' % time.strftime('%Y-%m-%d %H:%M:%S'))
        labels_adt, labels_kid = load_best_image_labels()
        gallery.build_sub_feature_map(0, labels_adt, labels_kid)  # 0.3指取前70%
        convert_method = gallery.convert_files_to_dataframe_avg
        print('\n[%s] labels' % time.strftime('%Y-%m-%d %H:%M:%S'))
        X = convert_method(['labels'])
    elif name == MODEL_OCRS:
        # for text features
        print('\n[%s] ocrs' % time.strftime('%Y-%m-%d %H:%M:%S'))
        X = learner_text.bert_client.encode(gallery.get_all_strings())
    if grid:
        train_models_grid(X, Y)
    else:
        train_default_models(X, Y)


def explain_model():
    import joblib, json
    market = 'huawei'
    name = 'image'
    model = joblib.load('%s/model_%s.pkl' % (market, name))
    if os.path.exists('%s/vals_%s.json' % (market, name)):
        with open('%s/vals_%s.json' % (market, name), "rb") as handle:
            val = json.load(handle)
    dataset = MyGallery(sub_features=val[0])
    X = dataset.convert_files_to_dataframe_avg(['labels'])
    explainer = LimeTabularExplainer(np.array(X), feature_names=X.columns, class_names=['nkid', 'kid'])
    for i, x in X.iterrows():
        print(dataset.file_list[i])
        exp = explainer.explain_instance(x, model.predict_proba, num_features=10)
        print('true label =', ['nkid', 'kid'][dataset.labels[i]])
        print('prob(kid) =', model.predict_proba(np.array(x).reshape(1, -1))[0, 1])
        for e in exp.as_list():
            print(e)


if __name__ == '__main__':
    # explain_model()
    # 对feature文件夹中每个文件，读取icon和screenshots的值，处理icon和每个screenshots（远程）和对应image文件夹下所有文件（本地）
    # label在2000个应用上调参的最好结合方式是ll union zx
    # for labels_adt, labels_kid in load_all_image_labels():
        # print(labels_adt)
        # print(labels_kid)
    labels_adt, labels_kid = load_best_image_labels()
    # print(len(labels_adt), len(labels_kid))
    gallery = MyGallery()
    print('kid:', gallery.pos_count, 'adt', gallery.neg_count)
    Y = gallery.labels

    feature_type = [MODEL_IMGS, MODEL_OCRS][0]
    if feature_type == MODEL_IMGS:
        # for img features
        print('\n[%s] counting features ...' % time.strftime('%Y-%m-%d %H:%M:%S'))
        gallery.build_sub_feature_map(0, labels_adt, labels_kid)  # 0.3指取前70%
        # gallery.build_sub_feature_map(0, None, None)  # 0.3指取前70%
        convert_method = gallery.convert_files_to_dataframe_avg

        print('\n[%s] labels' % time.strftime('%Y-%m-%d %H:%M:%S'))
        X = convert_method(['labels'])
        # select_best_chi2(X, Y, 0, feature_size_assigned[feature_type])
        # train_hard_grid(X, Y)  # 效果比机器学习差得多，舍弃
        train_default_models(X, Y)
        # train_models_grid(X, Y)
        #
        # print('\n[%s] avg hsv' % time.strftime('%Y-%m-%d %H:%M:%S'))
        # X = convert_method(['avg_hsv'])
        # train_default_models(X, Y)
        # # train_models(X, Y)

    # 不同类型的数据scale不同，可以分别试试不处理或做数据归一化后训练模型哪种效果更好

    elif feature_type == MODEL_OCRS:
        # for text features
        print('\n[%s] ocrs' % time.strftime('%Y-%m-%d %H:%M:%S'))
        X = learner_text.bert_client.encode(gallery.get_all_strings())
        # train_models_grid(X, Y)
        train_default_models(X, Y)

